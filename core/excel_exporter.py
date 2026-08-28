"""
Módulo para la exportación y conversión a archivos de Excel.
"""
import os
import datetime
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename

def exportar_diccionarios_excel(resultados, output_dir, acad_prompt=None):
    """Exporta los resultados escaneados desde AutoCAD a un archivo Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        if acad_prompt:
            acad_prompt("Error: openpyxl no instalado. Ejecuta: pip install openpyxl\n")
        else:
            print("Error: openpyxl no instalado.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Bloques Exportados"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["No.", "Nombre Bloque", "Tipo", "Coordenada X", "Coordenada Y"]
    all_attr_tags = set()
    for r in resultados:
        all_attr_tags.update(r["atributos"].keys())
    sorted_tags = sorted(all_attr_tags)
    headers.extend(sorted_tags)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    postes_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    tramos_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

    for idx, r in enumerate(resultados, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=r["nombre_bloque"]).border = thin_border
        ws.cell(row=row, column=3, value=r["tipo_detectado"]).border = thin_border
        ws.cell(row=row, column=4, value=r["x"]).border = thin_border
        ws.cell(row=row, column=5, value=r["y"]).border = thin_border

        row_fill = postes_fill if "CAT_POSTE" in r["tipo_detectado"].upper() else tramos_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = row_fill

        for tag_idx, tag in enumerate(sorted_tags):
            col = 6 + tag_idx
            val = r["atributos"].get(tag, "")
            ws.cell(row=row, column=col, value=val).border = thin_border

    auto_widths = {1: 6, 2: 22, 3: 16, 4: 14, 5: 14}
    for col, width in auto_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    for tag_idx, tag in enumerate(sorted_tags):
        col_letter = ws.cell(row=1, column=6 + tag_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(tag) + 4, 14)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bloques_exportados_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath

def convertir_csv_a_excel_interactivo(csv_path):
    """Convierte un archivo CSV a Excel procesando las columnas de atributos."""
    try:
        if not os.path.exists(csv_path):
            print(f"Error: No se encuentra el archivo {csv_path}")
            return

        df = pd.read_csv(csv_path, sep=",", quotechar='"', encoding="utf-8")

        if "Atributos" in df.columns:
            def extraer_atributos(texto):
                dicc = {}
                if pd.isna(texto) or not str(texto).strip():
                    return dicc
                pares = str(texto).split(";")
                for par in pares:
                    if "=" in par:
                        clave, valor = par.split("=", 1)
                        dicc[clave.strip()] = valor.strip()
                return dicc

            df_attr = df["Atributos"].apply(extraer_atributos).apply(pd.Series)
            df = pd.concat([df.drop(columns=["Atributos"]), df_attr], axis=1)

        Tk().withdraw()
        excel_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar archivo Excel como..."
        )
        if not excel_path:
            print("Operación cancelada por el usuario.")
            return

        df.to_excel(excel_path, index=False)
        print(f"Excel generado en: {excel_path}")
        
        # Usar os.startfile solo en windows
        if os.name == 'nt':
            os.startfile(excel_path)
        else:
            print(f"Por favor abra manualmente: {excel_path}")

    except Exception as e:
        print(f"Error en el proceso de conversión de Python: {str(e)}")
